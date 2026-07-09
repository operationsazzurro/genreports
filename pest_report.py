from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import datetime
import requests
from PIL import Image as PILImage
import base64
import tempfile
from io import BytesIO
import os
from concurrent.futures import ThreadPoolExecutor
import groupdocs_conversion_cloud

from groupdocs_conversion_cloud import ConvertApi
from groupdocs_conversion_cloud import ConvertDocumentDirectRequest


GROUPDOCS_CLIENT_ID = os.environ.get("GROUPDOCS_CLIENT_ID")
GROUPDOCS_CLIENT_SECRET = os.environ.get("GROUPDOCS_CLIENT_SECRET")

convert_api = groupdocs_conversion_cloud.ConvertApi.from_keys(
    GROUPDOCS_CLIENT_ID, GROUPDOCS_CLIENT_SECRET
)

print("CLIENT ID:", GROUPDOCS_CLIENT_ID)
print(
    "CLIENT SECRET LENGTH:",
    len(GROUPDOCS_CLIENT_SECRET) if GROUPDOCS_CLIENT_SECRET else "None",
)


def _download_and_validate(url, timeout=8):
    """Downloads a URL and confirms the bytes are a genuinely complete,
    decodable image — not just a 200 status with truncated/partial content.
    Returns raw bytes on success, None on any failure."""
    response = requests.get(url, timeout=timeout)
    if response.status_code != 200:
        print(f"[fetch] {url} -> non-200 status: {response.status_code}")
        return None

    content = response.content

    expected_len = response.headers.get("Content-Length")
    if expected_len is not None and len(content) != int(expected_len):
        print(
            f"[fetch] {url} -> truncated download: got {len(content)} bytes, "
            f"expected {expected_len}"
        )
        return None

    try:
        PILImage.open(BytesIO(content)).verify()
    except Exception as e:
        print(f"[fetch] {url} -> failed integrity check: {e}")
        return None

    return content


def _fetch_image_bytes(url, width=305, height=305, retries=2):
    """Download + resize a single image, with validation and retry."""
    if not url:
        return None

    last_error = None
    for attempt in range(1, retries + 1):
        try:
            content = _download_and_validate(url)
            if content is None:
                last_error = "failed validation or non-200 status"
                continue

            pil_img = PILImage.open(BytesIO(content)).convert("RGB")
            pil_img.thumbnail((width * 4, height * 4), PILImage.LANCZOS)

            converted = BytesIO()
            pil_img.save(converted, format="JPEG", quality=80, optimize=True)
            converted.seek(0)

            PILImage.open(BytesIO(converted.getvalue())).verify()
            converted.seek(0)
            return converted

        except Exception as e:
            last_error = f"{type(e).__name__}: {e}"

    print(f"[fetch] {url} -> giving up after {retries} attempts ({last_error})")
    return None


def pest_report_fn(data, report_format, is_cancelled=None):
    wb = Workbook()

    # ========== COVER PAGE ==========
    cover = wb.active
    cover.title = "Cover Page"

    try:
        azzurro_logo = XLImage("azzurro.png")
        azzurro_logo.width, azzurro_logo.height = 108, 24
        cover.add_image(azzurro_logo, "A2")
    except Exception as e:
        print(f"Logo load failed: {e}")

    try:
        client_logo = XLImage("imdaad.png")
        client_logo.width, client_logo.height = 64, 71
        cover.add_image(client_logo, "K1")
    except Exception as e:
        print(f"Logo load failed: {e}")

    try:
        mainclient_logo = XLImage("awqaf.png")
        mainclient_logo.width, mainclient_logo.height = 319, 124
        cover.add_image(mainclient_logo, "D8")
    except Exception as e:
        print(f"Logo load failed: {e}")

    cover.merge_cells("A17:K17")
    site_title = cover["A17"]
    site_title.value = "GAIAE AD PACKAGE 3 - MOSQUES"
    site_title.font = Font(size=15, bold=True, color="808080")
    site_title.alignment = Alignment(horizontal="center", vertical="center")

    cover.merge_cells("A21:K21")
    title = cover["A21"]
    title.value = "PEST CONTROL ACTIVITY REPORT"
    title.font = Font(size=20, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")

    cover.merge_cells("A23:K23")
    date_cell = cover["A23"]
    date_cell.value = f"Generated on: {datetime.date.today().strftime('%d-%b-%Y')}"
    date_cell.alignment = Alignment(horizontal="center")

    cover.paperSize = cover.PAPERSIZE_LETTER
    cover.page_setup.orientation = cover.ORIENTATION_PORTRAIT
    cover.page_margins.left = 0.5
    cover.page_margins.right = 0.5
    cover.page_margins.top = 0.5
    cover.page_margins.bottom = 0.5
    cover.page_setup.fitToWidth = 1
    cover.page_setup.fitToHeight = 0
    cover.sheet_properties.pageSetUpPr.fitToPage = True
    cover.page_setup.centerHorizontally = True
    cover.page_setup.centerVertically = True

    # ========== DATA SHEET ==========
    ws = wb.create_sheet("Weekly Report")

    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "Weekly Pest Control Activity Report"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(
        start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
    )
    ws.row_dimensions[1].height = 25

    headers = ["Finance Code", "Mosque Name", "Activity", "Before Photo", "After Photo"]
    ws.append(headers)

    #
    rows_per_page = 7  # page 1 total: 2 header rows + 5 data rows
    header_rows = 2
    max_col = 5
    header_height = 25
    data_row_height = 224
    data_rows_per_page = rows_per_page - header_rows  # 6 data rows per page

    def _apply_block_border(ws, start_row, end_row, start_col, end_col):
        """Draws a table-style border over a rectangular block of cells:
        thick on the outer perimeter, thin on every inner cell edge."""
        thick = Side(style="thick", color="000000")
        thin = Side(style="thin", color="000000")

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                left = thick if col == start_col else thin
                right = thick if col == end_col else thin
                top = thick if row == start_row else thin
                bottom = thick if row == end_row else thin
                ws.cell(row=row, column=col).border = Border(
                    left=left, right=right, top=top, bottom=bottom
                )

    total_data_rows = len(data) * 2  # 2 excel rows per data row (4 images each)
    pages = (total_data_rows // data_rows_per_page) + (
        1 if total_data_rows % data_rows_per_page else 0
    )
    pages = max(pages, 1)

    for page in range(pages):
        if page == 0:
            page_start_row = 1
            page_end_row = rows_per_page
        else:
            page_start_row = rows_per_page + (page - 1) * data_rows_per_page + 1
            page_end_row = page_start_row + data_rows_per_page - 1

        for r in range(page_start_row, page_end_row + 1):
            local_row_index = (r - page_start_row) + 1
            if page == 0 and local_row_index <= header_rows:
                ws.row_dimensions[r].height = header_height
            else:
                ws.row_dimensions[r].height = data_row_height

        _apply_block_border(ws, page_start_row, page_end_row, 1, max_col)

    ws.print_title_rows = f"$1:${header_rows}"

    # FIX: this loop used to unconditionally overwrite row 2's border with
    # thin on all sides, which wiped out the thick left/right outer edge
    # that _apply_block_border() had already drawn for row 2 (it's inside
    # the page-1 block, at the block's start_col/end_col). Now it only
    # thins the inner edges and preserves the outer thick left/right border
    # on the first/last column.
    thin_side = Side(style="thin", color="000000")
    thick_side = Side(style="thick", color="000000")

    for col_idx, col in enumerate(
        ws.iter_cols(min_row=2, max_row=2, min_col=1, max_col=len(headers)), start=1
    ):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(
                start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
            )
            left = thick_side if col_idx == 1 else thin_side
            right = thick_side if col_idx == len(headers) else thin_side
            cell.border = Border(
                left=left, right=right, top=thin_side, bottom=thin_side
            )

    column_widths = [18, 30, 40, 42, 42]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ===== 1. Gather every URL needed, per row, safely padded to 4 =====
    row_urls = []
    for row in data:
        urls = (row.get("before_file_url") or "").split(",")
        urls = [u.strip() for u in urls]
        urls += [""] * (4 - len(urls))
        row_urls.append(urls[:4])

    BATCH_SIZE = 50
    MAX_WORKERS = 12

    row_index = 3
    total_rows = len(data)
    overall_start = datetime.datetime.now()

    for batch_start in range(0, total_rows, BATCH_SIZE):
        if is_cancelled and is_cancelled():
            print(
                f"Cancellation detected — stopping after {batch_start} of {total_rows} rows"
            )
            return None
        batch_rows = data[batch_start : batch_start + BATCH_SIZE]
        batch_urls_groups = row_urls[batch_start : batch_start + BATCH_SIZE]
        flat_urls = [u for urls in batch_urls_groups for u in urls]

        batch_num = (batch_start // BATCH_SIZE) + 1
        print(
            f"[batch {batch_num}] fetching {len(flat_urls)} images "
            f"for rows {batch_start + 1}-{batch_start + len(batch_rows)} of {total_rows}..."
        )
        batch_start_time = datetime.datetime.now()

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            flat_images = list(executor.map(_fetch_image_bytes, flat_urls))

        batch_elapsed = (datetime.datetime.now() - batch_start_time).total_seconds()
        failed_count = sum(1 for img in flat_images if img is None)
        print(
            f"[batch {batch_num}] done in {batch_elapsed:.1f}s "
            f"({failed_count}/{len(flat_urls)} failed)"
        )

        images_per_row = [flat_images[i : i + 4] for i in range(0, len(flat_images), 4)]

        for row, imgs in zip(batch_rows, images_per_row):
            img1_bytes, img2_bytes, img3_bytes, img4_bytes = imgs

            ws.cell(row=row_index, column=1, value=row.get("fc", ""))
            ws.cell(row=row_index, column=2, value=row.get("name", ""))
            ws.cell(row=row_index, column=3, value=row.get("activity", ""))

            if img1_bytes:
                img = XLImage(img1_bytes)
                img.width, img.height = 305, 305
                ws.add_image(img, f"D{row_index}")
            else:
                ws.cell(row=row_index, column=4, value="Image load failed")

            if img2_bytes:
                img = XLImage(img2_bytes)
                img.width, img.height = 305, 305
                ws.add_image(img, f"E{row_index}")
            else:
                ws.cell(row=row_index, column=5, value="Image load failed")

            ws.cell(row=row_index + 1, column=1, value=row.get("fc", ""))
            ws.cell(row=row_index + 1, column=2, value=row.get("name", ""))
            ws.cell(row=row_index + 1, column=3, value=row.get("activity", ""))

            if img3_bytes:
                img = XLImage(img3_bytes)
                img.width, img.height = 305, 305
                ws.add_image(img, f"D{row_index + 1}")
            else:
                ws.cell(row=row_index + 1, column=4, value="Image load failed")

            if img4_bytes:
                img = XLImage(img4_bytes)
                img.width, img.height = 305, 305
                ws.add_image(img, f"E{row_index + 1}")
            else:
                ws.cell(row=row_index + 1, column=5, value="Image load failed")

            for i in range(1, 4):
                ws.cell(row=row_index, column=i).alignment = Alignment(
                    wrap_text=True, vertical="center", horizontal="center"
                )
                ws.cell(row=row_index + 1, column=i).alignment = Alignment(
                    wrap_text=True, vertical="center", horizontal="center"
                )

            ws.row_dimensions[row_index].height = 224
            ws.row_dimensions[row_index + 1].height = 224
            row_index += 2

    overall_elapsed = (datetime.datetime.now() - overall_start).total_seconds()
    print(f"All batches complete in {overall_elapsed:.1f}s")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_xlsx:
        wb.save(tmp_xlsx.name)
        tmp_xlsx_path = tmp_xlsx.name

    print("File exists:", os.path.exists(tmp_xlsx_path))
    print("File path:", tmp_xlsx_path)
    print("File Size:", os.path.getsize(tmp_xlsx_path))

    if report_format == "excel":
        return tmp_xlsx_path

    # ===== PDF path via GroupDocs =====
    if not GROUPDOCS_CLIENT_ID or not GROUPDOCS_CLIENT_SECRET:
        raise RuntimeError(
            "GROUPDOCS_CLIENT_ID / GROUPDOCS_CLIENT_SECRET environment "
            "variables are not set — PDF conversion cannot proceed."
        )

    try:
        request_conv = ConvertDocumentDirectRequest("pdf", tmp_xlsx_path)

        pdf_response = convert_api.convert_document_direct(request_conv)

    except Exception:
        import traceback

        traceback.print_exc()
        raise

    if isinstance(pdf_response, str) and os.path.exists(pdf_response):
        pdf_path = pdf_response
    elif isinstance(pdf_response, bytes):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
            tmp_pdf.write(pdf_response)
            pdf_path = tmp_pdf.name
    elif isinstance(pdf_response, str):
        try:
            pdf_bytes = base64.b64decode(pdf_response)
        except Exception:
            raise TypeError("Response string is not valid Base64 and not a path.")
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
            tmp_pdf.write(pdf_bytes)
            pdf_path = tmp_pdf.name
    else:
        raise TypeError(f"Unexpected response type: {type(pdf_response)}")

    print("PDF conversion successful:", pdf_path)
    return pdf_path
